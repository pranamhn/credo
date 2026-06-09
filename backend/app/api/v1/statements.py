"""
ENG-06 — Statement API endpoints (Dev mode: local file storage, sync parsing).
"""
from __future__ import annotations
import uuid
import calendar
from dataclasses import asdict
from datetime import date
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status, BackgroundTasks
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy import case, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.database import get_db
from app.models import Company, Statement, Transaction, RiskResult, AuditLog
from app.models.statement import StatementStatus
from app.schemas.statement import StatementRead, TransactionRead, TransactionPatch
from app.schemas.risk import RiskResultRead

router = APIRouter(prefix="/statements", tags=["statements"])

ALLOWED_EXTS = {".pdf", ".csv", ".xlsx"}
MAX_BYTES = settings.max_file_size_mb * 1024 * 1024
UPLOAD_DIR = Path("/tmp/risklens_uploads")
UPLOAD_DIR.mkdir(exist_ok=True)


def _period_month_to_date(period: str, end_of_month: bool = False) -> date | None:
    if not period:
        return None
    try:
        year, month = (int(part) for part in period.split("-", 1))
        day = calendar.monthrange(year, month)[1] if end_of_month else 1
        return date(year, month, day)
    except Exception:
        return None


def _count_pdf_pages(file_path: str) -> int | None:
    try:
        import pdfplumber

        with pdfplumber.open(file_path) as pdf:
            return len(pdf.pages)
    except Exception:
        return None


def _statement_page_count(stmt: Statement) -> int:
    parse_meta = stmt.parse_meta if isinstance(stmt.parse_meta, dict) else {}
    page_count = parse_meta.get("page_count")
    if page_count:
        return int(page_count)

    storage_key = getattr(stmt, "storage_key", None)
    if storage_key and Path(storage_key).suffix.lower() == ".pdf" and Path(storage_key).exists():
        return int(_count_pdf_pages(storage_key) or 0)

    return 0


async def _attach_statement_quality(db: AsyncSession, statements: list[Statement]) -> None:
    ids = [stmt.id for stmt in statements]
    if not ids:
        return

    txn_result = await db.execute(
        select(
            Transaction.statement_id,
            func.avg(Transaction.confidence).label("statement_confidence"),
            func.sum(case((Transaction.confidence < 1, 1), else_=0)).label("low_confidence_count"),
        )
        .where(Transaction.statement_id.in_(ids))
        .group_by(Transaction.statement_id)
    )
    txn_quality = {
        row.statement_id: {
            "statement_confidence": float(row.statement_confidence) if row.statement_confidence is not None else None,
            "low_confidence_count": int(row.low_confidence_count or 0),
        }
        for row in txn_result
    }

    risk_result = await db.execute(
        select(RiskResult.statement_id, RiskResult.flag_count).where(RiskResult.statement_id.in_(ids))
    )
    anomaly_counts = {row.statement_id: int(row.flag_count or 0) for row in risk_result}

    for stmt in statements:
        quality = txn_quality.get(stmt.id, {})
        setattr(stmt, "statement_confidence", quality.get("statement_confidence"))
        setattr(stmt, "low_confidence_count", quality.get("low_confidence_count", 0))
        setattr(stmt, "anomaly_count", anomaly_counts.get(stmt.id, 0))
        setattr(stmt, "page_count", _statement_page_count(stmt))


async def _run_parsing_background(statement_id: str, file_path: str) -> None:
    """Run parsing pipeline and persist results (called as background task)."""
    from sqlalchemy import create_engine
    from sqlalchemy.orm import Session
    from app.models.statement import StatementStatus
    from app.parsers.pipeline import ParsingPipeline
    from app.risk.metrics import compute_metrics
    from app.risk.flags import RedFlagEngine
    from decimal import Decimal

    engine = create_engine(settings.database_url_sync)
    with Session(engine) as db:
        stmt = db.get(Statement, uuid.UUID(statement_id))
        if not stmt:
            return

        stmt.status = StatementStatus.parsing
        db.commit()

        try:
            pipeline = ParsingPipeline()
            canonical = pipeline.run(file_path)

            metrics = compute_metrics(canonical)
            flag_engine = RedFlagEngine()
            flags = flag_engine.analyze(canonical, metrics.estimated_monthly_income)

            # Replace old results only after the new parse succeeds. This keeps
            # existing data intact if a re-parse fails for a bank-specific format.
            db.query(Transaction).filter(Transaction.statement_id == uuid.UUID(statement_id)).delete()
            db.query(RiskResult).filter(RiskResult.statement_id == uuid.UUID(statement_id)).delete()

            # Persist transactions
            for txn in canonical.transactions:
                db.add(Transaction(
                    statement_id=uuid.UUID(statement_id),
                    row=txn.row,
                    date=txn.date,
                    value_date=txn.value_date,
                    description_raw=txn.description_raw,
                    description_normalized=txn.description_normalized,
                    debit=txn.debit,
                    credit=txn.credit,
                    balance=txn.balance,
                    category=txn.category,
                    flags=[str(f) for f in (txn.flags or [])],
                    confidence=txn.confidence,
                    source=txn.source,
                    is_low_confidence=txn.is_low_confidence,
                    raw_meta=txn.raw_meta,
                ))

            # Persist risk result
            risk = RiskResult(
                statement_id=uuid.UUID(statement_id),
                total_credit=metrics.total_credit,
                total_debit=metrics.total_debit,
                net_flow=metrics.net_flow,
                avg_daily_balance=metrics.avg_daily_balance,
                min_balance=metrics.min_balance,
                max_balance=metrics.max_balance,
                days_below_threshold=metrics.days_below_threshold,
                negative_balance_days=metrics.negative_balance_days,
                transaction_count=metrics.transaction_count,
                estimated_monthly_income=metrics.estimated_monthly_income,
                estimated_monthly_obligations=metrics.estimated_monthly_obligations,
                dsr=metrics.dsr,
                flags={k: {
                    "flag_type": v.flag_type,
                    "severity": v.severity,
                    "count": v.count,
                    "total_amount": str(v.total_amount) if v.total_amount else None,
                    "supporting_rows": v.supporting_rows,
                    "description": v.description,
                    "confidence": v.confidence,
                } for k, v in flags.flags.items()},
                flag_count=flags.flag_count,
                has_judol=flags.has_judol,
                has_pinjol=flags.has_pinjol,
                has_passthrough=flags.has_passthrough,
                category_summary={k: str(v) for k, v in (metrics.category_summary or {}).items()},
            )
            db.add(risk)

            recon = canonical.reconciliation
            stmt.bank_code = canonical.bank_code
            stmt.bank_name = canonical.bank_name
            stmt.account_no_masked = canonical.account_no_masked
            stmt.account_holder = canonical.account_holder
            stmt.period_start = canonical.period_start
            stmt.period_end = canonical.period_end
            stmt.opening_balance = canonical.opening_balance
            stmt.closing_balance = canonical.closing_balance
            stmt.is_reconciled = recon.balanced if recon else False
            stmt.reconciliation_delta = recon.delta if recon else None
            stmt.detection_confidence = canonical.detection_confidence
            stmt.is_scanned = (canonical.parse_meta.ocr if canonical.parse_meta else False)
            existing_meta = stmt.parse_meta if isinstance(stmt.parse_meta, dict) else {}
            parsed_meta = canonical.parse_meta.model_dump() if canonical.parse_meta else {}
            stmt.parse_meta = {**existing_meta, **parsed_meta}
            stmt.status = StatementStatus.done if (recon and recon.balanced) else StatementStatus.needs_review
            stmt.parsed_at = datetime.utcnow()
            db.commit()

        except Exception as exc:
            stmt.status = StatementStatus.failed
            stmt.parse_error = str(exc)
            db.commit()
            raise


def _parse_financial_document(stmt: Statement, file_path: str) -> None:
    """Parse financial statement PDFs synchronously during upload/reparse."""
    if Path(file_path).suffix.lower() != ".pdf":
        stmt.status = StatementStatus.done
        stmt.parsed_at = datetime.utcnow()
        return

    try:
        existing_meta = stmt.parse_meta if isinstance(stmt.parse_meta, dict) else {}
        meta_update: dict = {}
        parsed_ok = False

        if stmt.document_type == "profit_loss":
            from app.parsers.pnl_parsing import parse_pnl_pdf

            report = parse_pnl_pdf(file_path)
            report_dict = asdict(report)
            stmt.period_start = _period_month_to_date(report.period_start)
            stmt.period_end = _period_month_to_date(report.period_end, end_of_month=True)
            stmt.currency = "IDR" if "rupiah" in report.currency.lower() else (report.currency or "IDR")
            meta_update = {
                "parser": "pnl_parsing",
                "pnl": report_dict,
                "summary": report.summaries,
            }
            parsed_ok = bool(report.line_items)

        elif stmt.document_type == "balance_sheet":
            from app.parsers.bs_parser import parse_balance_sheet_pdf

            report = parse_balance_sheet_pdf(file_path)
            report_dict = asdict(report)
            stmt.period_start = _period_month_to_date(report.period_start)
            stmt.period_end = _period_month_to_date(report.period_end, end_of_month=True)
            stmt.currency = "IDR" if "rupiah" in report.currency.lower() else (report.currency or "IDR")
            stmt.is_reconciled = bool(report.balance_checks) and all(check.balanced for check in report.balance_checks)
            stmt.reconciliation_delta = max((abs(check.delta or 0) for check in report.balance_checks), default=None)
            meta_update = {
                "parser": "bs_parser",
                "balance_sheet": report_dict,
                "summary": report.summaries,
            }
            parsed_ok = bool(report.line_items)

        elif stmt.document_type == "cash_flow":
            from app.parsers.cf_parser import parse_cash_flow_pdf

            report = parse_cash_flow_pdf(file_path)
            report_dict = asdict(report)
            stmt.period_start = _period_month_to_date(report.period_start)
            stmt.period_end = _period_month_to_date(report.period_end, end_of_month=True)
            stmt.currency = "IDR"
            stmt.opening_balance = report.cash_check.opening_cash
            stmt.closing_balance = report.cash_check.ending_cash
            stmt.is_reconciled = report.cash_check.balanced
            stmt.reconciliation_delta = report.cash_check.delta
            meta_update = {
                "parser": "cf_parser",
                "cash_flow": report_dict,
                "summary": report.summaries,
            }
            parsed_ok = bool(report.line_items)

        else:
            stmt.status = StatementStatus.done
            stmt.parsed_at = datetime.utcnow()
            return

        stmt.parse_meta = {**existing_meta, **meta_update}
        stmt.status = StatementStatus.done if parsed_ok else StatementStatus.needs_review
        stmt.parse_error = None
        stmt.parsed_at = datetime.utcnow()
    except Exception as exc:
        stmt.status = StatementStatus.failed
        stmt.parse_error = str(exc)
        stmt.parsed_at = datetime.utcnow()


@router.post("/upload", response_model=StatementRead, status_code=status.HTTP_202_ACCEPTED)
async def upload_statement(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    company_id: uuid.UUID | None = Form(None),
    document_type: str = Form("bank_statement"),
    db: AsyncSession = Depends(get_db),
):
    ext = Path(file.filename or "").suffix.lower()
    if ext not in ALLOWED_EXTS:
        raise HTTPException(400, f"Format tidak didukung: {ext}")

    content = await file.read()
    if len(content) > MAX_BYTES:
        raise HTTPException(413, f"File terlalu besar (maks {settings.max_file_size_mb}MB)")
    if company_id and not await db.get(Company, company_id):
        raise HTTPException(404, "Company not found")

    allowed_document_types = {"bank_statement", "profit_loss", "cash_flow", "balance_sheet", "other", "nib", "ahu", "akta"}
    if document_type not in allowed_document_types:
        raise HTTPException(400, f"Tipe dokumen tidak didukung: {document_type}")

    statement_id = uuid.uuid4()
    local_path = str(UPLOAD_DIR / f"{statement_id}{ext}")

    with open(local_path, "wb") as f:
        f.write(content)

    page_count = _count_pdf_pages(local_path) if ext == ".pdf" else None

    DEV_USER_ID = uuid.UUID("00000000-0000-0000-0000-000000000001")
    stmt = Statement(
        id=statement_id,
        uploaded_by=DEV_USER_ID,  # dev mode: fixed dev user
        company_id=company_id,
        document_type=document_type,
        original_filename=file.filename or "unknown",
        storage_key=local_path,
        file_size_bytes=len(content),
        mime_type=file.content_type,
        parse_meta={"page_count": page_count} if page_count is not None else None,
        status=StatementStatus.queued if document_type == "bank_statement" else StatementStatus.done,
    )
    if document_type in {"profit_loss", "balance_sheet", "cash_flow"}:
        _parse_financial_document(stmt, local_path)

    db.add(stmt)
    db.add(AuditLog(
        user_id=None,
        statement_id=statement_id,
        action="upload",
        detail={"filename": file.filename, "size_bytes": len(content), "company_id": str(company_id) if company_id else None, "document_type": document_type},
    ))
    await db.flush()
    await db.commit()

    # Parse bank statements in background. Financial statements are parsed
    # synchronously because they only extract positioned report text and do not
    # create txns.
    if document_type == "bank_statement":
        background_tasks.add_task(_run_parsing_background, str(statement_id), local_path)

    result = await db.get(Statement, statement_id)
    return result


@router.get("", response_model=list[StatementRead])
async def list_statements(
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Statement).order_by(Statement.created_at.desc()).limit(100)
    )
    statements = list(result.scalars().all())
    await _attach_statement_quality(db, statements)
    return statements


@router.get("/{statement_id}", response_model=StatementRead)
async def get_statement(
    statement_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
):
    stmt = await db.get(Statement, statement_id)
    if not stmt:
        raise HTTPException(404, "Statement not found")
    await _attach_statement_quality(db, [stmt])
    return stmt


@router.delete("/{statement_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_statement(
    statement_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
):
    stmt = await db.get(Statement, statement_id)
    if not stmt:
        raise HTTPException(404, "Statement not found")

    # Remove file from disk if it still exists
    storage_key = getattr(stmt, "storage_key", None)
    if storage_key:
        try:
            Path(storage_key).unlink(missing_ok=True)
        except Exception:
            pass  # Non-fatal — proceed with DB deletion

    await db.delete(stmt)
    await db.commit()


@router.post("/{statement_id}/reparse", response_model=StatementRead)
async def reparse_statement(
    statement_id: uuid.UUID,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
):
    """Re-trigger parsing for a failed or stuck statement."""
    stmt = await db.get(Statement, statement_id)
    if not stmt:
        raise HTTPException(404, "Statement not found")
    if stmt.status in (StatementStatus.queued, StatementStatus.parsing):
        raise HTTPException(409, "Statement sedang diproses")
    if stmt.document_type not in {"bank_statement", "profit_loss", "balance_sheet", "cash_flow"}:
        raise HTTPException(400, "Hanya bank statement, Profit & Loss, Balance Sheet, dan Cash Flow yang bisa di-reparse")

    stmt.status = StatementStatus.queued
    stmt.parse_error = None  # type: ignore[assignment]
    stmt.parsed_at = None  # type: ignore[assignment]

    if stmt.document_type in {"profit_loss", "balance_sheet", "cash_flow"}:
        _parse_financial_document(stmt, stmt.storage_key)
        await db.commit()
        await db.refresh(stmt)
        await _attach_statement_quality(db, [stmt])
        return stmt

    await db.commit()
    await db.refresh(stmt)

    background_tasks.add_task(_run_parsing_background, str(statement_id), stmt.storage_key)
    await _attach_statement_quality(db, [stmt])
    return stmt


@router.get("/{statement_id}/transactions", response_model=list[TransactionRead])
async def list_transactions(
    statement_id: uuid.UUID,
    skip: int = Query(0, ge=0),
    limit: int = Query(200, ge=1, le=5000),
    category: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
):
    q = select(Transaction).where(Transaction.statement_id == statement_id)
    if category:
        q = q.where(Transaction.category == category)
    q = q.order_by(Transaction.row).offset(skip).limit(limit)
    result = await db.execute(q)
    return result.scalars().all()


@router.patch("/{statement_id}/transactions/{row}", response_model=TransactionRead)
async def patch_transaction(
    statement_id: uuid.UUID,
    row: int,
    patch: TransactionPatch,
    db: AsyncSession = Depends(get_db),
):
    q = select(Transaction).where(
        Transaction.statement_id == statement_id,
        Transaction.row == row,
    )
    result = await db.execute(q)
    txn = result.scalar_one_or_none()
    if not txn:
        raise HTTPException(404, f"Row {row} tidak ditemukan")

    for field, value in patch.model_dump(exclude_unset=True).items():
        setattr(txn, field, value)
    txn.is_manually_corrected = True
    txn.is_low_confidence = False
    txn.confidence = 1.0
    txn.source = "manual"
    return txn


@router.get("/{statement_id}/risk", response_model=RiskResultRead)
async def get_risk(
    statement_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
):
    q = select(RiskResult).where(RiskResult.statement_id == statement_id)
    result = await db.execute(q)
    risk = result.scalar_one_or_none()
    if not risk:
        raise HTTPException(404, "Risk result belum tersedia")
    return risk


@router.get("/{statement_id}/export")
async def export_statement(
    statement_id: uuid.UUID,
    format: str = Query("xlsx", pattern="^(xlsx|json)$"),
    db: AsyncSession = Depends(get_db),
):
    stmt = await db.get(Statement, statement_id)
    if not stmt:
        raise HTTPException(404, "Statement not found")

    q = select(Transaction).where(Transaction.statement_id == statement_id).order_by(Transaction.row)
    result = await db.execute(q)
    txns = result.scalars().all()

    if format == "json":
        data = {
            "statement_id": str(statement_id),
            "bank_code": stmt.bank_code,
            "account_holder": stmt.account_holder,
            "period_start": str(stmt.period_start),
            "period_end": str(stmt.period_end),
            "opening_balance": float(stmt.opening_balance or 0),
            "closing_balance": float(stmt.closing_balance or 0),
            "is_reconciled": stmt.is_reconciled,
            "transactions": [
                {
                    "row": t.row, "date": str(t.date),
                    "description": t.description_raw,
                    "debit": float(t.debit or 0),
                    "credit": float(t.credit or 0),
                    "balance": float(t.balance or 0),
                    "category": t.category, "flags": t.flags or [],
                }
                for t in txns
            ],
        }
        return JSONResponse(content=data)

    # xlsx
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transaksi"
    headers = ["Row", "Tanggal", "Keterangan", "Debit", "Kredit", "Saldo", "Kategori", "Flag"]
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    white = Font(bold=True, color="FFFFFF")
    red_fill = PatternFill(fill_type="solid", fgColor="FFD7D7")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, t in enumerate(txns, 2):
        row_data = [t.row, str(t.date), t.description_raw,
                    float(t.debit or 0), float(t.credit or 0),
                    float(t.balance or 0), t.category or "", ", ".join(t.flags or [])]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=i, column=col, value=val)
        if t.flags:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = red_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"statement_{statement_id}_{stmt.bank_code}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
